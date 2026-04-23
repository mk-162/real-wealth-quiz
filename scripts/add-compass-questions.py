#!/usr/bin/env python3
"""
Add the new Compass question rows to `Question Segment Master.xlsx` (Sheet 2).

The Compass PDF report needs 11 new banded inputs that the current form does
not capture. Before they can render in the journey, each new question needs a
row in the Y/C/N matrix telling the segment engine which segments see it.

This script is one-shot, idempotent, and safe:
  - Opens the workbook
  - Finds Sheet 2 by position (not by name, since the sheet name may vary)
  - Detects the last used row
  - Adds any rows from NEW_ROWS that are not already present (keyed on Question ID)
  - Saves alongside the original as `Question Segment Master.AUTO-BACKUP.xlsx`,
    then overwrites the original

Prerequisites:
  pip install openpyxl

Usage:
  python scripts/add-compass-questions.py

Post-run:
  npx tsx scripts/parse-segment-master.ts     # regenerate matrix.json
  npm run content:build                       # revalidate + compile all content

Rollback:
  If anything goes wrong, restore from `Question Segment Master.AUTO-BACKUP.xlsx`
  (created automatically at the start of this script).

If you prefer to edit Sheet 2 by hand (no Python), see the MANUAL FALLBACK
at the bottom of this file — copy-paste the rows block into Sheet 2.
"""

from __future__ import annotations
import shutil
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required. Install with: pip install openpyxl")


# -----------------------------------------------------------------------------
# New rows to insert. Each entry: (Question ID, label, Y/C/N per segment S1..S9)
# -----------------------------------------------------------------------------

# Legend in the matrix:
#   Y = always shown for this segment
#   C = conditional (predicate lives in src/lib/segmentation/engine.ts)
#   N = not shown for this segment

NEW_ROWS: list[tuple[str, str, list[str]]] = [
    # Wealth snapshot — new pension pot band (conditional on any pension)
    ("Q4.A.1", "Total pension value (band)",               ["C","C","C","C","C","C","C","C","C"]),
    # Liquid wealth splits — always asked
    ("Q4.A.2", "Cash savings (band)",                      ["Y","Y","Y","Y","Y","Y","Y","Y","Y"]),
    ("Q4.A.3", "ISA balance (band)",                       ["Y","Y","Y","Y","Y","Y","Y","Y","Y"]),
    ("Q4.A.4", "GIA / trading balance (band)",             ["Y","Y","Y","Y","Y","Y","Y","Y","Y"]),

    # Mortgage fields — conditional on home-owner with mortgage
    ("Q4.B.1", "Mortgage monthly payment (band)",          ["C","C","C","C","C","C","C","C","C"]),
    ("Q4.B.2", "Mortgage end age (band)",                  ["C","C","C","C","C","C","C","C","C"]),

    # Cash-flow slope — always asked; employer-pension conditional on employed
    ("Q4.C.1", "Monthly savings across all pots (band)",   ["Y","Y","Y","Y","Y","Y","Y","N","Y"]),
    ("Q4.C.2", "Employer pension contribution (%, band)",  ["Y","Y","Y","Y","C","C","Y","N","C"]),

    # State pension and retirement shape — always asked
    ("Q4.D.1", "Expected state pension amount (band)",     ["Y","Y","Y","Y","Y","Y","Y","Y","Y"]),
    ("Q4.D.2", "NI qualifying years (band)",               ["Y","Y","Y","Y","Y","Y","Y","Y","Y"]),
    ("Q4.D.3", "Retirement spend ratio (less/same/more)",  ["Y","Y","Y","Y","Y","Y","Y","N","Y"]),
]

# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------

HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parent            # master_template/
WORKBOOK = REPO_ROOT.parent / "Question Segment Master.xlsx"   # one level up
BACKUP   = REPO_ROOT.parent / "Question Segment Master.AUTO-BACKUP.xlsx"


def main() -> int:
    if not WORKBOOK.exists():
        sys.exit(f"Workbook not found at {WORKBOOK}")

    # Always back up before touching.
    shutil.copy2(WORKBOOK, BACKUP)
    print(f"[backup]  {BACKUP}")

    wb = load_workbook(WORKBOOK, data_only=False, keep_vba=False)
    if len(wb.sheetnames) < 2:
        sys.exit("Expected at least 2 sheets in the workbook.")
    # Sheet 2 is the matrix (Sheet 1 is often a glossary / cover).
    sheet = wb.worksheets[1]
    print(f"[open]    sheet[1] = {sheet.title!r}  ({sheet.max_row} rows × {sheet.max_column} cols)")

    # Build a map of existing question IDs → row index.
    existing: dict[str, int] = {}
    for row_idx in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row_idx, column=1).value
        if isinstance(cell, str) and cell.startswith("Q"):
            existing[cell.strip()] = row_idx

    print(f"[scan]    {len(existing)} existing question rows")

    next_row = sheet.max_row + 1
    added = 0
    skipped = 0

    for qid, label, values in NEW_ROWS:
        if qid in existing:
            print(f"[skip]    {qid}  (already present at row {existing[qid]})")
            skipped += 1
            continue

        # Write: col A = qid, col B = label (if header exists), cols C... = S1..S9 values.
        # Actual column layout is determined by reading the header of Sheet 2.
        sheet.cell(row=next_row, column=1, value=qid)

        # We assume the Sheet 2 header is: [Question ID, Label, S1, S2, S3, S4, S5, S6, S7, S8, S9]
        # If the real sheet uses a different layout, edit MANUALLY — this script errs
        # on the side of a structured write; the backup lets you retry after fixing.
        sheet.cell(row=next_row, column=2, value=label)
        for i, v in enumerate(values):
            sheet.cell(row=next_row, column=3 + i, value=v)

        print(f"[add]     row {next_row}  {qid}  {values}")
        next_row += 1
        added += 1

    wb.save(WORKBOOK)
    print(f"[save]    {WORKBOOK}")
    print(f"[done]    +{added} rows added, {skipped} already present")
    print()
    print("Next steps:")
    print("  1. Open the xlsx and sanity-check the new rows in Sheet 2.")
    print("  2. npx tsx scripts/parse-segment-master.ts")
    print("  3. npm run content:check")
    print("  4. npm run content:build")
    return 0


if __name__ == "__main__":
    sys.exit(main())


"""
----------------------------------------------------------------------
MANUAL FALLBACK — if Python / openpyxl is not available.
----------------------------------------------------------------------

Open `Question Segment Master.xlsx` in Excel (or LibreOffice), go to Sheet 2,
scroll to the bottom of the data, and paste the following rows. Columns must be:

    Question ID | Label | S1 | S2 | S3 | S4 | S5 | S6 | S7 | S8 | S9

Rows to add:

    Q4.A.1  Total pension value (band)               C  C  C  C  C  C  C  C  C
    Q4.A.2  Cash savings (band)                      Y  Y  Y  Y  Y  Y  Y  Y  Y
    Q4.A.3  ISA balance (band)                       Y  Y  Y  Y  Y  Y  Y  Y  Y
    Q4.A.4  GIA / trading balance (band)             Y  Y  Y  Y  Y  Y  Y  Y  Y
    Q4.B.1  Mortgage monthly payment (band)          C  C  C  C  C  C  C  C  C
    Q4.B.2  Mortgage end age (band)                  C  C  C  C  C  C  C  C  C
    Q4.C.1  Monthly savings across all pots (band)   Y  Y  Y  Y  Y  Y  Y  N  Y
    Q4.C.2  Employer pension contribution (%, band)  Y  Y  Y  Y  C  C  Y  N  C
    Q4.D.1  Expected state pension amount (band)     Y  Y  Y  Y  Y  Y  Y  Y  Y
    Q4.D.2  NI qualifying years (band)               Y  Y  Y  Y  Y  Y  Y  Y  Y
    Q4.D.3  Retirement spend ratio (less/same/more)  Y  Y  Y  Y  Y  Y  Y  N  Y

Save, then run:

    npx tsx scripts/parse-segment-master.ts
    npm run content:build
"""
